
from openpyxl import Workbook

# 创建电子表格并初始化
book = Workbook()
sheet = book.create_sheet('pt_rest', 0)

sheet.cell(1, 2, 'EM')
sheet.cell(1, 3, 'F1')

with open('pt_rest.txt', 'r', encoding='utf-8') as f:
    lines = f.readlines()
    n = 2
    for line in lines:
        tokens = line.strip().split()
        if len(tokens) > 0:
            sheet.cell(n, 1, n - 1)
            sheet.cell(n, 2, tokens[-3])
            sheet.cell(n, 3, tokens[-1])
            n += 1

book.save('pt_rest.xlsx')
